import re
from openpyxl import load_workbook
from commons.getLog import Log
import os
import sys
sys.path.append(os.path.dirname(os.path.abspath(__file__)))
from interface_test import InterfaceTest
import json

'''
本class：
遍历课运行的所有case:读取excel字段，调用interface_test
每个case调用intercase_test
参数关联
'''




class RunTestCase:
    def __init__(self):
        global log
        log=Log()

    
    #获取、执行用例
    def run(self,testcase_file,report_file=None):
        #对case处理
        testcase_file=os.path.join(os.getcwd(),testcase_file)
        if not os.path.exists(testcase_file):
            log.error('testcase_file,测试用例文件不存在，去哪了')
            sys.exit()

        wb=load_workbook(testcase_file)
        table=wb['TestCase']#选择第二个sheet:TestCase
        print('--table----max---',table.max_row)


        #预留字典，为了后续处理接口关联
        correlationDict={}


        #遍历所有case
        for i in range(2,table.max_row+1):
            #判断用例,执行状态是No的跳过
            if str(table.cell(row=i,column=10).value.replace('\n','').replace('\r',''))=='No':
                continue

            #获取excel中数据,拿掉换行、回车
            num=str(int(table.cell(row=i,column=1).value)).replace('\n','').replace('\r','')
            print('--编号--',num)           
            api_purpose=table.cell(row=i,column=2).value.replace('\n','').replace('\r','')
            print('--接口名称--',api_purpose)          
            api_host=table.cell(row=i,column=3).value.replace('\n','').replace('\r','')
            # print('--地址前缀--',api_host)          
            request_url=table.cell(row=i,column=4).value.replace('\n','').replace('\r','')
            # print('--请求地址--',request_url)          
            request_method=table.cell(row=i,column=5).value.replace('\n','').replace('\r','')
            # print('--请求方法--',request_method)          
            request_data_type=table.cell(row=i,column=6).value.replace('\n','').replace('\r','')
            # print('--请求格式--',request_data_type)       
            request_data=table.cell(row=i,column=7).value.replace('\n','').replace('\r','')
            print('--请求数据--',request_data)      
            check_point=table.cell(row=i,column=8).value.replace('\n','').replace('\r','')
            # print('--检查点--',chect_point)
            correlation=table.cell(row=i,column=9).value
            print('--关联参数--',correlation)

        #     #如果把请求参数发给到了txt中，则读取里面的内容。这里注意txt编码必须utf-8无bom格式的
        #     if os.path.exists(request_data):
        #         fopen=open(request_data,encoding='utf-8')
        #         request_data=fopen.readline()
        #         fopen.close()

            '''
            在request_data中查找是否存在需要关联的请求数据
            关联参数的处理在后面进行
            '''
            for keyword in correlationDict:
                if request_data.find(keyword)>0:
                    request_data=request_data.replace(keyword,str(correlationDict[keyword]))

            #将准备好的所有数据传入下面的方法进行接口测试
            it=InterfaceTest()
            status,response=it.interface_test(num,api_purpose,api_host,request_url,request_data,check_point,request_method,request_data_type,i,table,log)
       
            #关联参数处理
            if correlation !=None:
                correlation=correlation.replace('\n','').replace('\r','').split(';')

                for j in range(len(correlation)):
                    #根据=把关联数据拆分
                    param = correlation[j].split('=')
                    print("param=",param)
                    print("param[0]=",param[0])
                    print("param[1]=", param[1])
                    if len(param)==2:
                        if param[1]==''or not re.search(r'^\[',param[1]) :#or not re.rearch()
                            log.error(num+' '+api_purpose+'关联参数设置有误，请检查')
                            continue
                        # value=resp
                        for key in param[1][1:-1].split(']['):#对第三个数据进行拆分，从下表1开始到末尾
                            # print("key=",key)#key= checkstatus
                            # print('key类型--',type(key))
                            # print('response--',response)
                            temp = response[key]
                            print("temp=",temp)
                            #因为中间处理的时候数据会有变化，所以在给一个新的值存储
                            response = temp
                            print('response=',response)

                        #关联到的响应放到字典里，方便后续去遍历替换参数
                        correlationDict[param[0]] = response
                        print("correlationDict[param[0]]=", response)

                    else:
                        print("error")

    
        #save the file
        wb.save(testcase_file)
            
     
            # #save the file
            # wb.save()

            # #切换操作表格
            # # ws=table.create_sheet('testsheet01')
            # # ws1=table.active
            # # ws2=table['testsheet']







'''
import unittest
import json
import requests

import os
import sys

# 相对路径的import
# sys.path.append("../")
# 绝对路径的import
# sys.path.append(os.path.dirname(os.path.abspath(__file__)) + "/../")这个也可以，下面的也成
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from commons.getResponse import HttpRequestResponse

class MyTest(unittest.TestCase):
    def test_m1(self):
        url = "http://localhost:12306"
        path = "/login1"

        full_url = url+path
        print("POST请求完整url=", full_url)
        # headers = {
        #     "Content-Type": "application/x-www-form-urlencoded"
        # }
        data = {
            "username": "xiaoming",
            "pwd": "123456"
        }

        print("POST请求参数=", data)

        new_PostJson=HttpRequestResponse()
        form_json,status_code,time = new_PostJson.post_form(full_url, data=data)
        if status_code==200:
            self.assertEqual('successed', form_json['reason'], "响应不符合预期")
            print(path,"接口的响应时间=", time, '秒')
        else:
            print('响应状态非200')

    def test_m2(self):
        url = "http://localhost:12306"
        path = "/login2"

        full_url = url+path
        print("POST请求完整url=", full_url)
        # headers = {
        #     "Content-Type": "application/json"
        # }
        data = {
            "username": "xiaoqiang",
            "pwd": "123456"
        }
        print("POST请求参数=", data)
        new_PostJson=HttpRequestResponse()
        json_json,status_code,time = new_PostJson.post_json(full_url, data=data)
        if status_code==200:
            self.assertEqual('successed', json_json['reason'], "响应不符合预期")
            print(path,"接口的响应时间=", time, '秒')
        else:
            print('响应状态非200')

    def test_m3(self):
        # get
        new_get = HttpRequestResponse()
        url = "http://localhost:12306"
        path = "/book_info"

        full_url = url+path
        print("GET请求完整url=", full_url)

        params = {
            "bookname": "接口来自moco",
            "checkstatus": "on"
        }
        get_json,status_code,time= new_get.get(full_url, params=params)
        print("get_json---", get_json)
        if status_code==200:
            self.assertEqual('successed', get_json['reason'], "响应不符合预期")
            print(path,"接口的响应时间=", time, '秒')
        else:
            print('响应状态非200')
'''