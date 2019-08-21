import re
import os
import sys
import json
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from commons.getResponse import HttpRequestResponse
import json
from openpyxl.styles import Font,colors,Alignment

'''
本class：
调用请求类发送请求，接收响应
根据响应结果进行判断
对比结果写入excel
'''


class InterfaceTest:
    def interface_test(self,num,api_purpose,api_host,request_url,
    request_data,check_point,request_method,request_data_type,i,table,log):
        full_request_url=api_host+request_url
        print('full_request_url啥类型',type(full_request_url))
        #注意从表里取出来的都是str类型，要用eval字符串转成字典
        # request_data=eval(request_data)

        http=HttpRequestResponse()
        if request_method=='POST'and request_data_type=='FORM':
            response,status_code,time=http.post_form(url=full_request_url,data=request_data)
        elif request_method=='GET':
            response,status_code,time=http.get(url=full_request_url,params=request_data)
        elif request_method=='POST'and request_data_type=='JSON':
            response,status_code,time=http.post_json(url=full_request_url,data=request_data)
        else:
            log.error(num+''+api_purpose+'请求方法出错，请确认请求方法字段是否有误')
            return 400,request_method

        '''
        注意，之前处理的把json类型转换成python类型进行，
        处理完成，为了展示效果，转换成json类型，使用方法json.dumps()
        其中参数ensure_ascii=False是为了解决中文编码
        '''

        response_json=json.dumps(response,ensure_ascii=False)
        if status_code==200:
            # re.search 扫描整个字符串并返回第一个成功的匹配。
            # re.search(pattern, string, flags=0)
            # pattern	匹配的正则表达式；string	要匹配的字符串。
            if re.search(check_point,response_json):
                table.cell(row=i,column=11).value='成功'
                table.cell(rew=i,column=12).value=response_json
                table.cell(row=i,column=13).value=time 
                log.info('编号'+num+' '+'接口名称='+api_purpose+'成功'+str(status_code)+' '+'响应时间'+str(time)+'秒'+'\n')
                return status_code,response
            else:
                myfont=Font(color=color.RED)
                table.cell(row=i,column=11).value='失败'
                #如果失败是红色字体
                table.cell(row=i,column=11).Font=myfont
                table.cell(row=i,column=12).value=response_json
                log.error(num+' '+api_purpose+',失败！，['+str(status_code))
                return 2001,response

        else:
            table.cell(row=i,column=11).value='失败'
            table.cell(row=i,column=12).value=response_json
            log.error(num+' '+api_purpose+',失败！，['+str(status_code)+']')
            return status_code,response


